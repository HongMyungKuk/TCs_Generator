import pandas as pd
import openpyxl as oxl
import math

class Parser : 
    def __init__(self) :
        pass

    def Init(self, filepath) :
        self.__scenario_id = ''
        self.Load(filepath)

    def Load(self, filepath) :
        self.__wb  = oxl.load_workbook(filepath)
        # self.__ws = self.__wb.active # 현재 활성화 된 시트
        self.__id_to_sheetname = self.__wb.sheetnames
        self.__sheetname_to_id = {}
        
        self.__SetSheetNmaeToID()

    def __CalcSkipValue(self, filepath, sheetname) :
        skipvalue = 0
        sucess_flag = False

        data = pd.read_excel(filepath, sheet_name=sheetname, usecols='B')
        df = pd.DataFrame(data)

        row_size = len(df)
        col_size = len(df.columns)

        for row in range(0, row_size) : 
            row_B = str(df.iloc[row, 0])
            skipvalue += 1
            if row_B == '시나리오 ID' or row_B == 'ID' :
                sucess_flag = True 
                self.__scenario_id = row_B
                print('Find SkipValue [Success]\n')
                return skipvalue
        
        if sucess_flag == False :
            skipvalue = -1

        return skipvalue

    def LoadSheetData(self, filepath, sheetname, skipvalue) :
        skipvalue = self.__CalcSkipValue(filepath, sheetname)
        if skipvalue == -1 : 
            print("Find SkipValue [Failed]\n")
            exit()

        # usecol 파라미터 정책 필요!
        # 각 vux sheet 마다 col의 갯수가 변동됨
        # ****GUI 입력 vs 사전처리*****
        self.__data = pd.read_excel(filepath, sheet_name=sheetname, skiprows=skipvalue)
        old_df = pd.DataFrame(self.__data)
        old_df = old_df.iloc[:,1:]

        # col 이름 불러들이기
        # 함수 수정 필요!!
        col_names = self.Test_GetColmnName(old_df)

        new_df = old_df[col_names]
        self.__df = new_df
        self.__SetSize()

    def ProcessingData(self) :
        self.__scenario_id_data = []

        first_scenario_id = str(self.__df[self.__scenario_id][0])
        index = first_scenario_id.find('_')
        # '_' 문자가 나오기전까지의 문자열을 반환한다.
        scenario_id_key = str(self.__df[self.__scenario_id][0][0 : index])
        print(scenario_id_key)

        cnt = 0
        for i in range(0, self.__row_size) :
            scenario_id = self.__df[self.__scenario_id][i]
            if 'nan' == str(scenario_id) or 'NaN' == str(scenario_id):
                self.__df = self.__df.drop(i, axis=0)
                continue
            if scenario_id_key in scenario_id :
                print(str(cnt) + '.' + str(scenario_id))
                self.__scenario_id_data.append(str(scenario_id))
                cnt += 1
                continue
            # 시나리오 ID 값이 아닌것들은 데이터 프레임에서 제거한다.
            else : 
                self.__df = self.__df.drop(i, axis=0)

        # 모든 행이 Nan이면 그 행은 제거한다.
        self.__df = self.__df.dropna(how='all')
        self.__SetSize()

    def GetAllSenarioID(self) :
        return self.__scenario_id_data

    def __SetSheetNmaeToID(self) :
        i = 0
        for ele in self.__id_to_sheetname:
            self.__sheetname_to_id[ele] = i
            i += 1 

    def __SetSize(self) :
        self.__row_size = len(self.__df)
        self.__col_size = len(self.__df.columns)

    def SetSize(self, width, height) :
        self.__row_size = width
        self.__col_size = height

    def GetSize(self) : 
        return self.__row_size, self.__col_size
    
    def GetData(self) :
        return self.__df
    
    def Test_GetColmnName(self, data_frame) : 
        return list(data_frame.columns)
    
    def GetColmnName(self) :
        return list(self.__df.columns)
    
    def GetSheetNames(self) :
        return self.__id_to_sheetname

    def PrintSheetData(self, sheetname) :
        if sheetname == self.__id_to_sheetname[self.__sheetname_to_id[sheetname]] :
            sheet = self.__wb[sheetname]
            print(self.__wb[sheetname])

            for row in range(0, sheet.max_row) :
                cols = []
                for col in sheet.iter_cols(min_col=0, max_col=sheet.max_column):
                    cols.append(col[row].value)
                print(cols)
        else : 
            print('find not sheetname')


